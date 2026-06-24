from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime
import io, secrets
import pandas as pd

from database import get_db
import models, schemas
from auth import get_current_user, require_admin, hash_password, log_action

router = APIRouter(prefix="/api/v1/users", tags=["users"])


@router.get("")
def list_users(db: Session = Depends(get_db),
               current_user: models.User = Depends(require_admin)):
    users = db.query(models.User).order_by(models.User.created_at.desc()).all()
    result = []
    for u in users:
        emp = db.query(models.Employee).filter(models.Employee.id == u.employee_id).first() if u.employee_id else None
        # หาโครงการจากทุก source
        proj_names = []

        # SUP → หาจาก Project.sup_user_id
        if u.role == "sup":
            sup_projs = db.query(models.Project).filter(
                models.Project.sup_user_id == u.id
            ).all()
            proj_names = [p.name for p in sup_projs if p.name]

        elif emp:
            # Assignment (Admin assign)
            assigns = db.query(models.Assignment).filter(
                models.Assignment.employee_id == emp.id,
                models.Assignment.is_active == True
            ).all()
            assign_proj_ids = set()
            for a in assigns:
                p = db.query(models.Project).filter(models.Project.id == a.project_id).first()
                if p and p.name not in proj_names:
                    proj_names.append(p.name)
                    assign_proj_ids.add(p.id)
            # SupTeamMember (SUP assign)
            team_rows = db.query(models.SupTeamMember).filter(
                models.SupTeamMember.employee_id == emp.id
            ).all()
            for t in team_rows:
                if t.project_id not in assign_proj_ids:
                    p = db.query(models.Project).filter(models.Project.id == t.project_id).first()
                    if p and p.name not in proj_names:
                        proj_names.append(p.name)

        proj_name = ", ".join(proj_names) if proj_names else None
        result.append({
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "employee_id": u.employee_id,
            "is_active": u.is_active,
            "last_login": u.last_login.isoformat() if u.last_login else None,
            "created_at": u.created_at.isoformat() if u.created_at else None,
            "employee_name": f"{emp.first_name} {emp.last_name}" if emp else None,
            "photo_url": emp.photo_url if emp else None,
            "project_name": proj_name,
        })
    return result


@router.post("", response_model=schemas.UserOut, status_code=201)
def create_user(body: schemas.UserCreate,
                db: Session = Depends(get_db),
                current_user: models.User = Depends(require_admin)):
    existing = db.query(models.User).filter(models.User.username == body.username).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Username '{body.username}' มีอยู่แล้ว")

    user = models.User(
        username=body.username,
        role=body.role,
        employee_id=body.employee_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    log_action(db, current_user, "CREATE", "users", user.id,
               f"เพิ่มผู้ใช้: {user.username} (role: {user.role})")

    out = schemas.UserOut.model_validate(user)
    if user.employee_id:
        emp = db.query(models.Employee).filter(models.Employee.id == user.employee_id).first()
        out.employee_name = f"{emp.first_name} {emp.last_name}" if emp else None
    return out


@router.put("/{user_id}", response_model=schemas.UserOut)
def update_user(user_id: int, body: schemas.UserUpdate,
                db: Session = Depends(get_db),
                current_user: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    if user.username == "admin" and body.role and body.role != "admin":
        raise HTTPException(status_code=400, detail="ไม่สามารถเปลี่ยน role ของ admin หลักได้")

    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(user, field, value)
    user.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(user)

    log_action(db, current_user, "UPDATE", "users", user_id,
               f"แก้ไขผู้ใช้: {user.username}")
    out = schemas.UserOut.model_validate(user)
    if user.employee_id:
        emp = db.query(models.Employee).filter(models.Employee.id == user.employee_id).first()
        out.employee_name = f"{emp.first_name} {emp.last_name}" if emp else None
    return out


@router.post("/{user_id}/reset-password")
def reset_password(user_id: int, body: dict,
                   db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_admin)):
    """Admin reset password ให้ user — ถ้าไม่ส่ง new_password จะ generate auto"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    if user.username == "admin" and current_user.username != "admin":
        raise HTTPException(status_code=403, detail="ไม่มีสิทธิ์ reset password admin")

    new_pw = body.get("new_password", "").strip()
    if not new_pw:
        new_pw = secrets.token_urlsafe(10)  # auto-generate

    user.hashed_password = hash_password(new_pw)
    user.updated_at = datetime.utcnow()
    db.commit()

    log_action(db, current_user, "UPDATE", "users", user_id,
               f"Reset password: {user.username}")
    return {"success": True, "username": user.username, "new_password": new_pw}


@router.post("/import/excel")
def import_users(file: UploadFile = File(...),
                 db: Session = Depends(get_db),
                 current_user: models.User = Depends(require_admin)):
    """
    Import ผู้ใช้งานจาก Excel
    คอลัมน์: username (บังคับ), role (admin/sup/employee), employee_code (optional)
    - ถ้า username มีอยู่แล้ว → update role/employee_id (upsert)
    - password จะถูก set เป็น temp random → admin ส่ง Invite Link เพื่อให้ user ตั้งรหัสใหม่
    """
    try:
        contents = file.file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="ไม่สามารถอ่านไฟล์ Excel ได้")

    df.columns = [str(c).strip() for c in df.columns]

    if "username" not in df.columns and "ชื่อผู้ใช้" not in df.columns:
        raise HTTPException(status_code=400, detail="ไม่พบคอลัมน์ 'username' หรือ 'ชื่อผู้ใช้' ในไฟล์")

    valid_roles = {"admin", "sup", "employee"}
    created, updated, errors = 0, 0, []

    for idx, row in df.iterrows():
        username = str(row.get("username") or row.get("ชื่อผู้ใช้") or "").strip()
        if not username or username == "nan":
            continue

        sp = db.begin_nested()  # savepoint — rollback เฉพาะแถวนี้
        try:
            role_raw = str(row.get("role") or row.get("สิทธิ์") or "employee").strip().lower()
            role = role_raw if role_raw in valid_roles else "employee"

            # หา employee_id จาก employee_code ถ้ามี
            employee_id = None
            emp_code = str(row.get("employee_code") or row.get("รหัสพนักงาน") or "").strip()
            if emp_code and emp_code != "nan":
                emp = db.query(models.Employee).filter(
                    models.Employee.employee_code == emp_code
                ).first()
                if emp:
                    employee_id = emp.id

            existing = db.query(models.User).filter(models.User.username == username).first()

            if existing:
                # upsert — update role และ employee_id
                existing.role = role
                if employee_id is not None:
                    existing.employee_id = employee_id
                existing.updated_at = datetime.utcnow()
                updated += 1
            else:
                # สร้างใหม่ด้วย temp password (user ต้องใช้ Invite Link เพื่อตั้งรหัส)
                temp_pw = secrets.token_urlsafe(12)
                user = models.User(
                    username=username,
                    hashed_password=hash_password(temp_pw),
                    role=role,
                    employee_id=employee_id,
                    is_active=True,
                )
                db.add(user)
                db.flush()
                created += 1

            sp.commit()

        except Exception as e:
            sp.rollback()  # rollback เฉพาะ savepoint นี้ — แถวก่อนหน้าไม่หาย
            errors.append(f"แถว {idx+2} ({username}): {str(e)[:120]}")
            continue

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"บันทึกฐานข้อมูลล้มเหลว: {str(e)}")
    msg = f"นำเข้าสำเร็จ: เพิ่มใหม่ {created} บัญชี, อัพเดต {updated} บัญชี"
    if errors:
        msg += f" | ข้าม {len(errors)} แถว"
    log_action(db, current_user, "IMPORT", "users", description=msg)
    return {"success": True, "message": msg, "errors": errors,
            "note": "ผู้ใช้ใหม่ได้รับ temp password — กรุณาส่ง Invite Link เพื่อให้ตั้งรหัสใหม่"}




@router.get("/export/template")
def export_user_template(current_user: models.User = Depends(require_admin)):
    """Export Excel template สำหรับ import users"""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["username", "role", "employee_code"]
    header_notes = ["ชื่อผู้ใช้ (บังคับ)", "admin / sup / employee", "รหัสพนักงาน (optional)"]

    hdr_fill = PatternFill("solid", fgColor="4F46E5")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, (h, note) in enumerate(zip(headers, header_notes), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col, value=note).font = Font(italic=True, color="888888")

    # sample rows
    samples = [
        ("john.doe", "employee", "EMP001"),
        ("jane.sup", "sup", "EMP002"),
    ]
    for r, row in enumerate(samples, 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_import_template.xlsx"})

@router.delete("/{user_id}")
def deactivate_user(user_id: int, db: Session = Depends(get_db),
                    current_user: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="ไม่สามารถปิดการใช้งาน admin หลักได้")
    user.is_active = False
    db.commit()
    log_action(db, current_user, "DELETE", "users", user_id,
               f"ปิดการใช้งาน: {user.username}")
    return {"success": True, "message": "ปิดการใช้งานผู้ใช้แล้ว"}


# ── เปลี่ยน Role ──────────────────────────────────────

@router.patch("/{user_id}/role")
def change_user_role(user_id: int, body: dict,
                     db: Session = Depends(get_db),
                     current_user: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    new_role = body.get("role", "")
    if new_role not in ("employee", "sup", "admin"):
        raise HTTPException(status_code=400, detail="role ต้องเป็น employee / sup / admin")
    old_role = user.role
    user.role = new_role
    user.updated_at = datetime.utcnow()
    db.commit()
    log_action(db, current_user, "UPDATE", "users", user_id,
               f"เปลี่ยน role: {old_role} -> {new_role} ({user.username})")
    return {"success": True, "role": new_role}


# ── เปลี่ยน Username ──────────────────────────────────

@router.patch("/{user_id}/username")
def change_username(user_id: int, body: dict,
                    db: Session = Depends(get_db),
                    current_user: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    new_username = str(body.get("username", "")).strip()
    if not new_username:
        raise HTTPException(status_code=400, detail="ระบุ username")
    exists = db.query(models.User).filter(
        models.User.username == new_username,
        models.User.id != user_id
    ).first()
    if exists:
        raise HTTPException(status_code=409, detail="username นี้ถูกใช้งานแล้ว")
    old = user.username
    user.username = new_username
    user.updated_at = datetime.utcnow()
    db.commit()
    log_action(db, current_user, "UPDATE", "users", user_id,
               f"เปลี่ยน username: {old} -> {new_username}")
    return {"success": True}


# ── Reset Password (Admin ตั้งรหัสให้) ───────────────

@router.patch("/{user_id}/password")
def reset_password(user_id: int, body: dict,
                   db: Session = Depends(get_db),
                   current_user: models.User = Depends(require_admin)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")
    new_pw = str(body.get("password", "")).strip()
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร")
    user.hashed_password = hash_password(new_pw)
    user.updated_at = datetime.utcnow()
    db.commit()
    log_action(db, current_user, "UPDATE", "users", user_id,
               f"reset password: {user.username}")
    return {"success": True}


# ── สร้าง Admin ใหม่ ──────────────────────────────────

@router.post("/create-admin", status_code=201)
def create_admin_user(body: dict,
                      db: Session = Depends(get_db),
                      current_user: models.User = Depends(require_admin)):
    username = str(body.get("username", "")).strip()
    password = str(body.get("password", "")).strip()
    if not username or not password:
        raise HTTPException(status_code=400, detail="ระบุ username และ password")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร")
    exists = db.query(models.User).filter(models.User.username == username).first()
    if exists:
        raise HTTPException(status_code=409, detail="username นี้ถูกใช้งานแล้ว")
    admin = models.User(
        username=username,
        hashed_password=hash_password(password),
        role="admin",
        is_active=True,
    )
    db.add(admin)
    db.commit()
    db.refresh(admin)
    log_action(db, current_user, "CREATE", "users", admin.id,
               f"สร้าง admin ใหม่: {username}")
    return {"success": True, "id": admin.id, "username": admin.username, "role": "admin"}


@router.post("/{user_id}/invite")
def generate_invite(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_admin),
):
    """สร้าง invite token สำหรับพนักงาน (Admin only) — token ใช้ได้ 48 ชั่วโมง"""
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="ไม่พบผู้ใช้")

    token = secrets.token_urlsafe(32)
    key = f"invite_token_{user_id}"
    existing = db.query(models.Setting).filter(models.Setting.key == key).first()
    if existing:
        existing.value = token
    else:
        db.add(models.Setting(key=key, value=token))
    db.commit()

    log_action(db, current_user, "INVITE", "users", user_id,
               f"สร้าง invite link สำหรับ {user.username}")
    return {
        "token": token,
        "username": user.username,
        "user_id": user_id,
    }
