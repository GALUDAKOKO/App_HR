from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
import io

import models
from database import get_db
from auth import get_current_user

router = APIRouter(prefix="/api/v1/holidays", tags=["holidays"])

# ─── helpers ─────────────────────────────────────────────────────────────────
WEEKDAY_NAMES = ["จันทร์","อังคาร","พุธ","พฤหัส","ศุกร์","เสาร์","อาทิตย์"]

def _is_holiday(db: Session, check_date_str: str) -> dict:
    """Return {is_holiday, name, holiday_type} for a date string YYYY-MM-DD."""
    try:
        d = datetime.strptime(check_date_str, "%Y-%m-%d").date()
    except ValueError:
        return {"is_holiday": False, "name": None, "holiday_type": None}

    # public holiday exact match
    pub = db.query(models.Holiday).filter(
        models.Holiday.date == check_date_str,
        models.Holiday.holiday_type == "public",
        models.Holiday.is_active == True,
    ).first()
    if pub:
        return {"is_holiday": True, "name": pub.name, "holiday_type": "public"}

    # weekly holiday — weekday 0=Mon…6=Sun, Python date.weekday() same
    weekday = d.weekday()
    weekly = db.query(models.Holiday).filter(
        models.Holiday.holiday_type == "weekly",
        models.Holiday.date == str(weekday),
        models.Holiday.is_active == True,
    ).first()
    if weekly:
        return {"is_holiday": True, "name": weekly.name, "holiday_type": "weekly"}

    return {"is_holiday": False, "name": None, "holiday_type": None}

# ─── GET /api/v1/holidays ─────────────────────────────────────────────────────
@router.get("")
def list_holidays(
    year: int = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if year is None:
        year = datetime.utcnow().year

    # Weekly config rows (date field stores weekday 0-6 as string)
    weekly = db.query(models.Holiday).filter(
        models.Holiday.holiday_type == "weekly",
        models.Holiday.is_active == True,
    ).all()

    # Public holidays for the requested year
    public_rows = db.query(models.Holiday).filter(
        models.Holiday.holiday_type == "public",
        models.Holiday.is_active == True,
        func.substr(models.Holiday.date, 1, 4) == str(year),
    ).order_by(models.Holiday.date).all()

    return {
        "weekly": [
            {"weekday": int(h.date), "name": h.name, "label": WEEKDAY_NAMES[int(h.date)]}
            for h in weekly
        ],
        "public": [
            {"id": h.id, "date": h.date, "name": h.name}
            for h in public_rows
        ],
        "year": year,
    }

# ─── GET /api/v1/holidays/check ───────────────────────────────────────────────
@router.get("/check")
def check_date(
    date: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return _is_holiday(db, date)

# ─── GET /api/v1/holidays/range ───────────────────────────────────────────────
@router.get("/range")
def holidays_in_range(
    start: str,
    end: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Return list of holiday dates between start and end (inclusive)."""
    try:
        s = datetime.strptime(start, "%Y-%m-%d").date()
        e = datetime.strptime(end, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format (YYYY-MM-DD)")

    weekly_set = set(
        int(h.date) for h in
        db.query(models.Holiday).filter(
            models.Holiday.holiday_type == "weekly",
            models.Holiday.is_active == True,
        ).all()
    )
    public_set = set(
        h.date for h in
        db.query(models.Holiday).filter(
            models.Holiday.holiday_type == "public",
            models.Holiday.is_active == True,
            models.Holiday.date >= start,
            models.Holiday.date <= end,
        ).all()
    )

    result = []
    cur = s
    from datetime import timedelta
    while cur <= e:
        ds = cur.strftime("%Y-%m-%d")
        if ds in public_set or cur.weekday() in weekly_set:
            result.append(ds)
        cur += timedelta(days=1)
    return result

# ─── POST /api/v1/holidays/weekly ─────────────────────────────────────────────
@router.post("/weekly")
def set_weekly(
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """body = {weekdays: [0,1,...6]}  — replaces weekly config entirely."""
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")

    weekdays = body.get("weekdays", [])

    # remove old weekly rows
    db.query(models.Holiday).filter(
        models.Holiday.holiday_type == "weekly"
    ).delete()

    for wd in weekdays:
        if 0 <= wd <= 6:
            row = models.Holiday(
                date=str(wd),
                name=WEEKDAY_NAMES[wd],
                holiday_type="weekly",
                is_active=True,
            )
            db.add(row)
    db.commit()
    return {"ok": True, "weekdays": weekdays}

# ─── POST /api/v1/holidays/public ─────────────────────────────────────────────
@router.post("/public")
def add_public(
    body: dict,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")

    date_str = body.get("date", "").strip()
    name = body.get("name", "").strip()
    if not date_str or not name:
        raise HTTPException(400, "date and name required")

    existing = db.query(models.Holiday).filter(
        models.Holiday.date == date_str,
        models.Holiday.holiday_type == "public",
    ).first()
    if existing:
        existing.name = name
        existing.is_active = True
        db.commit()
        return {"id": existing.id, "date": existing.date, "name": existing.name}

    row = models.Holiday(date=date_str, name=name, holiday_type="public", is_active=True)
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"id": row.id, "date": row.date, "name": row.name}

# ─── DELETE /api/v1/holidays/public/{id} ──────────────────────────────────────
@router.delete("/public/{holiday_id}")
def delete_public(
    holiday_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")

    row = db.query(models.Holiday).filter(
        models.Holiday.id == holiday_id,
        models.Holiday.holiday_type == "public",
    ).first()
    if not row:
        raise HTTPException(404, "Not found")
    db.delete(row)
    db.commit()
    return {"ok": True}

# ─── POST /api/v1/holidays/import ─────────────────────────────────────────────
@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(403, "Admin only")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — pip install openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    added = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        raw_date = row[0]
        name = str(row[1]).strip() if len(row) > 1 and row[1] else "วันหยุด"

        # Accept datetime or string
        if hasattr(raw_date, "strftime"):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            ds = str(raw_date).strip()
            # Try common Thai/ISO formats
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    date_str = datetime.strptime(ds, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            else:
                errors.append(f"Row {idx}: invalid date '{ds}'")
                continue

        existing = db.query(models.Holiday).filter(
            models.Holiday.date == date_str,
            models.Holiday.holiday_type == "public",
        ).first()
        if existing:
            existing.name = name
            existing.is_active = True
        else:
            db.add(models.Holiday(date=date_str, name=name, holiday_type="public", is_active=True))
        added += 1

    db.commit()
    return {"added": added, "errors": errors}


# ─── GET /api/v1/holidays/export ──────────────────────────────────────────────
@router.get("/export")
def export_holidays(
    year: int = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Export public holidays as Excel template (date, name columns)."""
    if year is None:
        year = datetime.utcnow().year

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import StreamingResponse
        import io
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    rows = db.query(models.Holiday).filter(
        models.Holiday.holiday_type == "public",
        models.Holiday.is_active == True,
        func.substr(models.Holiday.date, 1, 4) == str(year),
    ).order_by(models.Holiday.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"วันหยุด {year}"

    # Styles
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    alt_fill = PatternFill("solid", fgColor="EEF2FF")

    # Header
    ws.append(["วันที่ (YYYY-MM-DD)", "ชื่อวันหยุด"])
    for col in range(1, 3):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[1].height = 24

    # Data
    THAI_DAYS = ["จันทร์","อังคาร","พุธ","พฤหัสบดี","ศุกร์","เสาร์","อาทิตย์"]
    for i, h in enumerate(rows, start=2):
        try:
            wd = datetime.strptime(h.date, "%Y-%m-%d").weekday()
            day_label = THAI_DAYS[wd]
        except Exception:
            day_label = ""
        ws.append([h.date, h.name])
        fill = red_fill if i % 2 == 0 else alt_fill
        for col in range(1, 3):
            cell = ws.cell(i, col)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            cell.fill = fill
        ws.row_dimensions[i].height = 20

    # If no data, add sample rows
    if not rows:
        samples = [
            (f"{year}-01-01", "วันขึ้นปีใหม่"),
            (f"{year}-04-06", "วันจักรี"),
            (f"{year}-04-13", "วันสงกรานต์"),
            (f"{year}-04-14", "วันสงกรานต์"),
            (f"{year}-04-15", "วันสงกรานต์"),
            (f"{year}-05-01", "วันแรงงานแห่งชาติ"),
            (f"{year}-05-05", "วันฉัตรมงคล"),
            (f"{year}-06-03", "วันเฉลิมพระชนมพรรษา สมเด็จพระราชินี"),
            (f"{year}-07-28", "วันเฉลิมพระชนมพรรษา ร.10"),
            (f"{year}-08-12", "วันแม่แห่งชาติ"),
            (f"{year}-10-13", "วันนวมินทรมหาราช"),
            (f"{year}-10-23", "วันปิยมหาราช"),
            (f"{year}-12-05", "วันพ่อแห่งชาติ"),
            (f"{year}-12-10", "วันรัฐธรรมนูญ"),
            (f"{year}-12-31", "วันสิ้นปี"),
        ]
        for i, (d, n) in enumerate(samples, start=2):
            ws.append([d, n])
            fill = red_fill if i % 2 == 0 else alt_fill
            for col in range(1, 3):
                cell = ws.cell(i, col)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                cell.fill = fill
            ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"holidays_{year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

