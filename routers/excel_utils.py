"""
excel_utils.py — Pandas-free Excel helper (openpyxl only)
Replaces pandas for all Excel read/write in APP_HR_Render.
Works on any Python version without compilation.
"""
from openpyxl import Workbook, load_workbook
import io


def notna(val) -> bool:
    """Equivalent to pd.notna() — True if value is not None / NaN / empty string."""
    if val is None:
        return False
    s = str(val).strip()
    return s not in ("", "nan", "None", "NaN", "NaT")


def _write_sheet(ws, rows: list) -> None:
    """Write list-of-dicts to an openpyxl worksheet."""
    if not rows:
        ws.append(["ไม่มีข้อมูล"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def write_excel(buf: io.BytesIO, rows: list, sheet_name: str = "Sheet1") -> None:
    """Write single-sheet Excel to buffer (replaces pd.DataFrame + pd.ExcelWriter)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit
    _write_sheet(ws, rows)
    wb.save(buf)


def write_excel_multi(buf: io.BytesIO, sheets: list) -> None:
    """Write multi-sheet Excel to buffer.
    sheets = [("SheetName", rows_list), ...]
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    for sheet_name, rows in sheets:
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        _write_sheet(ws, rows)
    wb.save(buf)


def read_excel(file_bytes: bytes) -> list:
    """Read first sheet of Excel file, return list of dicts.
    Replaces pd.read_excel() + df.iterrows().
    Column names are stripped of whitespace.
    Empty rows (all None) are skipped automatically.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    # Build headers from first row
    headers = [
        str(h).strip() if (h is not None and str(h).strip() not in ("", "None")) else f"_col_{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # skip blank rows
        result.append(dict(zip(headers, row)))
    return result


def excel_columns(rows: list) -> list:
    """Return column names from first row of read_excel result."""
    if not rows:
        return []
    return list(rows[0].keys())
